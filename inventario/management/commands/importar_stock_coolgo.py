"""
Management Command: importar_stock_coolgo
-----------------------------------------------
Importa el archivo Excel de stock de Cool & Go hacia la base de datos.
Lee las columnas: Bodega, Zona, Ubicación, CodigoArticulo, Descripcion,
Disponible, Lote, FechaFabricacion, FechaVencimiento, EstadoUbicacion.

Uso:
    python manage.py importar_stock_coolgo
    python manage.py importar_stock_coolgo --archivo ruta/al/archivo.xlsx
"""
import os
from datetime import date, datetime

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.conf import settings

from inventario.models import Bodega, Zona, Ubicacion, Producto, Lote


class Command(BaseCommand):
    help = 'Importa el stock inicial de Cool & Go desde el archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            type=str,
            default=None,
            help='Ruta al archivo Excel (por defecto busca stcok_ejemplo_coolandgo.xlsx en BASE_DIR/..)'
        )

    def handle(self, *args, **options):
        # Determinar ruta del archivo
        archivo = options['archivo']
        if not archivo:
            archivo = os.path.join(
                settings.BASE_DIR.parent, 'stcok_ejemplo_coolandgo.xlsx'
            )

        if not os.path.exists(archivo):
            raise CommandError(
                f'No se encontró el archivo: {archivo}\n'
                f'Usa --archivo para especificar la ruta correcta.'
            )

        self.stdout.write(f'📂 Leyendo archivo: {archivo}')

        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active

        # Leer encabezados dinámicamente para ser resiliente a cambios en el Excel
        headers = [cell.value for cell in next(ws.iter_rows())]
        self.stdout.write(f'📋 Columnas encontradas: {headers}')

        creados = {'bodegas': 0, 'zonas': 0, 'ubicaciones': 0, 'productos': 0, 'lotes': 0}
        errores = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Mapear la fila al diccionario de columnas
            datos = dict(zip(headers, row))

            # Saltear filas vacías
            if not any(datos.values()):
                continue

            try:
                nombre_bodega = str(datos.get('Bodega', '') or '').strip()
                nombre_zona = str(datos.get('Zona', '') or '').strip()
                codigo_ubicacion = str(datos.get('Ubicación', '') or '').strip()
                codigo_producto = str(datos.get('CodigoArticulo', '') or '').strip()
                descripcion = str(datos.get('DescripcionArticulo', datos.get('Descripcion', '') or '') or '').strip()
                disponible = datos.get('Cantidad', 0) or 0
                numero_lote = str(datos.get('Serie/Lote', '') or '').strip()
                estado_ubic = str(datos.get('Estado Ubicacion', 'DISPONIBLE') or 'DISPONIBLE').strip().upper()

                # Parsear fechas con fallback a hoy si no existen en el Excel
                def parse_fecha(val):
                    if isinstance(val, (date, datetime)):
                        return val.date() if isinstance(val, datetime) else val
                    if isinstance(val, str) and val.strip():
                        try:
                            return datetime.strptime(val.strip(), '%d/%m/%Y').date()
                        except ValueError:
                            pass
                    return date.today()

                fecha_fab = parse_fecha(None)  # El Excel no tiene columna de fabricacion
                fecha_venc = parse_fecha(datos.get('Fecha Vecto.'))

                if not nombre_bodega or not codigo_producto:
                    continue

                # -------------------------------------------------------
                # DRY: get_or_create evita duplicados y reutiliza instancias
                # -------------------------------------------------------
                bodega, creada = Bodega.objects.get_or_create(nombre=nombre_bodega)
                if creada:
                    creados['bodegas'] += 1

                if nombre_zona:
                    zona, creada = Zona.objects.get_or_create(bodega=bodega, nombre=nombre_zona)
                    if creada:
                        creados['zonas'] += 1

                    if codigo_ubicacion:
                        # Mapear estado de ubicación del Excel
                        estado_map = {'DISPONIBLE': 'DISPONIBLE', 'OCUPADA': 'OCUPADA', 'BLOQUEADA': 'BLOQUEADA'}
                        estado_ubic_val = estado_map.get(estado_ubic, 'DISPONIBLE')
                        ubicacion, creada = Ubicacion.objects.get_or_create(
                            zona=zona, codigo=codigo_ubicacion,
                            defaults={'estado': estado_ubic_val}
                        )
                        if creada:
                            creados['ubicaciones'] += 1
                    else:
                        ubicacion = None
                else:
                    zona = None
                    ubicacion = None

                # Producto/SKU
                producto, creado = Producto.objects.get_or_create(
                    codigo=codigo_producto,
                    defaults={'nombre': descripcion or codigo_producto}
                )
                if creado:
                    creados['productos'] += 1

                # Lote - Solo si tiene número de lote o cantidad disponible
                if numero_lote or int(disponible) > 0:
                    lote_num = numero_lote or f'AUTO-{codigo_producto}-F{i}'
                    lote, creado = Lote.objects.get_or_create(
                        producto=producto,
                        numero_lote=lote_num,
                        defaults={
                            'ubicacion': ubicacion,
                            'cantidad_disponible': int(disponible),
                            'fecha_fabricacion': fecha_fab,
                            'fecha_vencimiento': fecha_venc,
                        }
                    )
                    if creado:
                        creados['lotes'] += 1

            except Exception as e:
                errores.append(f'Fila {i}: {e}')

        # -------------------------------------------------------
        # Resumen final
        # -------------------------------------------------------
        self.stdout.write(self.style.SUCCESS('\n✅ Importación completada:'))
        for k, v in creados.items():
            self.stdout.write(f'   → {k.capitalize()} creados: {v}')

        if errores:
            self.stdout.write(self.style.WARNING(f'\n⚠ Errores encontrados ({len(errores)}):'))
            for e in errores:
                self.stdout.write(f'   {e}')
